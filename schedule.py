from enum import IntEnum
from re import search
from openpyxl import load_workbook

"""
Если ячейка является склеенной, вызов cell.value вернет None -
нужно найти левый верхний угол склеенной ячейки, в которой будет 
хранится настоящее значение
"""
def read_value(cell):
    sheet = cell.parent
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value

"""
Обычно единственное значащее число в ячейке окружено текстом -
курс, группа, номер аудитории вне пристройки
"""
def first_number(value):
    if value is None:
        return 0
    match = search(r'\d+', value)
    if match:
        return match.group()
    return value

"""
Номер аудитории состоит из числа и, возможно, одной буквы - пристройка
"""
def auditory(value) -> int:
    if value is not None:
        match = search(r'\d+[а-яА-Яa-zA-Z]?', value)
        if match:
            return match.group()
    return 0

"""
Проводится ли пара по числителю, знаменателю или каждую неделю
"""
class Week_even(IntEnum):
    odd = False,
    even = True,
    both = 2,

class Time_column(IntEnum):
    Weekday = 1,
    Time = 2,

class Group_row(IntEnum):
    Course = 1,
    Number = 2,
    Specialty = 3,
    Sub_Specialty = 4,

"""
Метка, по которой можно определить, что пара
проводится дистанционно, и не имеет аудитории
"""
DISTANT_TAG = 'ДО'

class Lab:
    """
    Кто-то решил для красоты вставлять пустые строчки между днями недели,
    поэтому нужно корректировать четность номера ячейки (числитель-знаменатель)
    по четным дням недели
    """
    def __set_even(self, even: Week_even):
        self.even = even
        if self.weekday in ['Вторник', 'Четверг', 'Суббота'] and self.even != Week_even.both:
            self.even = not self.even

    def __init__(self,  lecturer: str, value: str, course: str, group: str, specialty: str, weekday: str, time: str, even: Week_even):
        self.lecturer = lecturer
        self.weekday = weekday
        self.distant = DISTANT_TAG in value
        self.course = first_number(course)
        self.group = first_number(group)
        if self.distant:
            self.room = None
        else:
            self.room = auditory(value)

        self.__set_even(even)
        self.specialty = specialty
        self.time = time.replace(' ', '')

    def __str__(self):
        if(self.distant):
            return f"  [{DISTANT_TAG}] ({self.time}): [{self.specialty}] {self.course}({self.group})"
        return f"  ({self.time}): {self.room}, [{self.specialty}] {self.course}({self.group})"

"""
Группировка пар по преподавателю. Распределяет пары по числителю-знаменателю
"""
class Schedule:

    """
    Где-то в таблице ячейки склеивают горизонтально, если пара сразу у 
    двух подгрупп - где-то нет. Логика непонятна, поэтому просто склеиваю
    пары с одним временем. Если пары в списке будут идти в случайном порядке, 
    не по расписанию, то метод перестанет работать.
    """
    def __same_lab(list, lab):
        if list != []:
            last = list[-1]
            if last.weekday == lab.weekday and last.time == lab.time:
                return True
        return False

    def __init__(self, lecturer: str, lab_list: list[Lab]):
        self.lecturer = lecturer
        self.even = []
        self.odd = []
        for lab in lab_list:
            if lab.even == Week_even.even or lab.even == Week_even.both:
                if not Schedule.__same_lab(self.even, lab):
                    self.even.append(lab)

        for lab in lab_list:
            if lab.even == Week_even.odd or lab.even == Week_even.both:
                if not Schedule.__same_lab(self.odd, lab):
                    self.odd.append(lab)


    """
    Пока не группирую нигде по дням недели пары, поэтому просто 
    последовательно проверяю текущий день недели с предыдущим
    """
    def __print_by_weekday(labs):
        if len(labs) == 0:
            return
        prev_weekday = labs[0].weekday
        print(prev_weekday)
        print(labs[0])
        for lab in labs[1:]:
            if lab.weekday != prev_weekday:
                print(lab.weekday)
            prev_weekday = lab.weekday
            print(lab)

    def print(self):
        print(f'Расписание {self.lecturer}')
        print('Числитель:')
        Schedule.__print_by_weekday(self.odd)

        print()
        print('Знаменатель:')
        Schedule.__print_by_weekday(self.even)


def is_merged_vertically_cell(ws, x: int, y: int) -> bool:
    cell = ws.cell(y, x)
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return merged_range.min_row !=  merged_range.max_row
    return False

"""
Если ячейки соединены вертикально, это значит, что
пара проводится каждую неделю
"""
def even_odd(ws, x: int, y: int) -> Week_even:
    value = Week_even.both
    if not is_merged_vertically_cell(ws, x, y):
        value = y % 2 == 0
    return value

"""
Первоначальные данные (аудитория, предмет) получаются из ячейки, 
все остальное из соответствующих крайних ячеек - время (слева) и группа (сверху)
"""
def parse_cell(ws, x: int, y: int, value: str, name: str) -> Lab:
    return Lab(name, 
               value,
               read_value(ws.cell(Group_row.Course, x)),
               read_value(ws.cell(Group_row.Number, x)),
               read_value(ws.cell(Group_row.Specialty, x)),
               read_value(ws.cell(y, Time_column.Weekday)),
               read_value(ws.cell(y, Time_column.Time)),
               even_odd(ws, x, y))

def parse_table(ws, name: str) -> Schedule:
    lab_list = []
    for y in range(Group_row.Sub_Specialty + 1, ws.max_row):
        for x in range(Time_column.Time + 1, ws.max_column):
            value = ws.cell(y, x).value
            if value is not None and name in value:
                lab_list.append(parse_cell(ws, x, y, value, name))
    return Schedule(name, lab_list)

if __name__ == "__main__":
    wb = load_workbook('schedule.xlsx')
    ws = wb.active
    # parse_table(ws, 'Соколов').print()
    parse_table(ws, 'Точилин').print()
        